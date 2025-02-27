import os
from dotenv import load_dotenv
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import NoAlertPresentException
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.alert import Alert
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver import Keys
from openpyxl import load_workbook
from selenium import webdriver
from time import sleep

file_alunos = load_workbook("alunos.xlsx")
abaMain = file_alunos["Folha1"]
lastRow = abaMain.max_row

def colandoProntAluno(alunosCadastrados, lastRow, abaMain):
    for linha, aluno in zip(range(2, lastRow+1), alunosCadastrados):
        _nome = abaMain.cell(linha, column=1).value
        seiPront = abaMain.cell(linha, column=6)
        pront = abaMain.cell(linha, column=7)
        if not _nome:
            break
        if _nome == aluno['nome']:
            seiPront.value = aluno['sei']
            pront.value = aluno['codigo']
        print(_nome, seiPront.value, pront.value)
    file_alunos.save("alunos.xlsx")
    file_alunos.close()
    print('planilha salva, com novos dados')

def colaAluno(aluno, lastRow, abaMain):
    for linha in range(2, lastRow+1):
        _nome = abaMain.cell(linha, column=1).value
        seiPront = abaMain.cell(linha, column=6)
        pront = abaMain.cell(linha, column=7)
        if not _nome:
            break
        if _nome == aluno['nome']:
            seiPront.value = aluno['sei']
            pront.value = aluno['codigo']
    print(aluno['nome'], aluno['sei'], aluno['codigo'])
    print(aluno['nome'], 'foi cadastrado')
    file_alunos.save("alunos.xlsx")
    file_alunos.close()

def logar(driver, user, passw):
    # Logando no SEI
    # User
    driver.find_element(By.ID, 'txtUsuario').send_keys(user)
    sleep(1)
    # Password
    driver.find_element(By.ID, 'pwdSenha').send_keys(passw)
    sleep(1)
    # dropdown
    findCPS = Select(driver.find_element(By.ID, 'selOrgao'))
    findCPS.select_by_visible_text("CEETEPS")
    sleep(1)
    # Click no botão
    driver.find_element(By.ID, 'sbmAcessar').click()
    sleep(1)

def copyProntuario(driver):
    driver.switch_to.frame(0)
    copyPron = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CLASS_NAME, "infraArvoreNoSelecionado"))
    )
    texto = copyPron.text
    return texto

def aceitarAlertGoogle(driver):
    try:
        alertaPresent = WebDriverWait(driver, 5).until(EC.alert_is_present())
        if alertaPresent:
            alert = Alert(driver)
            alert.accept()
    except:
        print("Nenhum alerta apareceu.")

def cadastroAluno(driver, lista_alunos): # , lista_alunos):
    for aluno in lista_alunos:
        # Armazenar a janela atual em uma varíavel
        driver.find_element(By.XPATH, "//*[@id='infraMenu']/li[10]/a/span").click() # Iniciar Processo
        sleep(2)
        
        driver.find_element(By.XPATH, "//*[@id='tblTipoProcedimento']/tbody/tr[4]/td/a[2]").click() # Prontuário de Aluno
        sleep(2)
        
        input_nome = driver.find_element(By.ID, "txtInteressadoProcedimento") # Interessados
        input_nome.send_keys(aluno["nome"])
        sleep(2)
        input_nome.send_keys(Keys.ENTER)
        sleep(2)
        #Aceitando o Alerta do Google
        aceitarAlertGoogle(driver)
        sleep(1)
        driver.find_element(By.ID, "txtDescricao").send_keys(aluno["codigo"]) # Especificação
        sleep(1)
        driver.find_element(By.ID, "divOptRestrito").click() # Nível de Acesso - Restrito

        dropDown_hipotese = WebDriverWait(driver, 3).until(
            EC.presence_of_element_located((By.ID, "selHipoteseLegal"))
        )
        select = Select(dropDown_hipotese)

        # Informação Pessoal
        sleep(3)
        pront = os.getenv("PRONT")
        select.select_by_visible_text(pront)
        sleep(1)
        driver.find_element(By.ID, "btnSalvar").click()
        sleep(1)
        
        # Copiando Prontuário
        textoProntuario = copyProntuario(driver)
        aluno["sei"] = textoProntuario
        
        driver.switch_to.default_content()
        # Voltando a page home
        elemento = driver.find_element(By.ID, "lnkControleProcessos")    
        # Pegar o valor do atributo onclick
        onclick_value = elemento.get_attribute("onclick")
        url = onclick_value.split("location.href='")[1].split("'")[0]
        
        urlHome = "https://sei.sp.gov.br/sei" + url

        colaAluno(aluno, lastRow, abaMain)

        driver.get(urlHome)
        sleep(2)
    driver.quit()
    return lista_alunos

load_dotenv()
user = os.getenv("USER")
passw = os.getenv("PASS")
site = os.getenv("SITE")
def criarProntuario(lista_alunos):
    opt = webdriver.ChromeOptions()
    opt.add_argument("--headless=new")
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=opt)
    driver.maximize_window()

    print(user, passw, site)
    driver.get(site)
    
    logar(driver, user, passw)
    # Cadastrando Prontuário de Aluno
    sleep(5)
    cadastroAluno(driver, lista_alunos) #, lista_alunos)
    # colandoProntAluno(alunosCadastrados, lastRow, abaMain)