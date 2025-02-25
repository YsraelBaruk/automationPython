from openpyxl import load_workbook
from Create import criarProntuario

def colandoProntAluno(_nome, lista, pront, lastRow, abaMain):
    for linha, aluno in zip(range(2, lastRow+1), lista):
        _nome = abaMain.cell(linha, column=1).value
        pront = abaMain.cell(linha, column=7)
        if not _nome:
            break
        if _nome == aluno['nome']:
            pront.value = aluno['codigo']   
        print(_nome, pront.value)
    file_alunos.save("alunos.xlsx")
    file_alunos.close()
    print('planilha salva, com novos dados')
file_alunos = load_workbook("alunos.xlsx")
aba_main = file_alunos["Folha1"]
lastRow = aba_main.max_row

lista_alunos = []
lista_prontuarios = []

for linha in range(2, lastRow+1):
    _nome = aba_main.cell(linha, column=1).value
    _ra = aba_main.cell(linha, column=2).value
    _curso = aba_main.cell(linha, column=4).value
    seiPront = aba_main.cell(linha, column=6).value
    
    if not _nome:
        break

    pront_aluno = _nome + " - " + "RA " + str(_ra) + " - " + _curso.upper()
    lista_alunos.append({"nome": _nome, "ra": _ra, "curso": _curso, "codigo": pront_aluno, "sei": None, "codigo": pront_aluno})
# colandoProntAluno(_nome, lista_alunos, pront_aluno, lastRow, aba_main)

alunosCadastrados = criarProntuario(lista_alunos)